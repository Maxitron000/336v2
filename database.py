import sqlite3
import pandas as pd
import openpyxl
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple
from config import DB_NAME, MAIN_ADMIN_ID, EXPORT_FILENAME

class Database:
    def __init__(self):
        self.db_name = DB_NAME
        self.init_database()
    
    def init_database(self):
        """–ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏—è –±–∞–∑—ã –¥–∞–Ω–Ω—ã—Ö"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        
        # –¢–∞–±–ª–∏—Ü–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY,
                username TEXT UNIQUE,
                full_name TEXT NOT NULL,
                is_admin BOOLEAN DEFAULT FALSE,
                status TEXT DEFAULT '–≤_—á–∞—Å—Ç–∏',
                last_location TEXT,
                last_status_change TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # –î–æ–±–∞–≤–ª—è–µ–º –Ω–æ–≤—ã–µ –∫–æ–ª–æ–Ω–∫–∏ –µ—Å–ª–∏ –∏—Ö –Ω–µ—Ç (–º–∏–≥—Ä–∞—Ü–∏—è)
        try:
            cursor.execute('ALTER TABLE users ADD COLUMN status TEXT DEFAULT "–≤_—á–∞—Å—Ç–∏"')
        except:
            pass  # –ö–æ–ª–æ–Ω–∫–∞ —É–∂–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç
        
        try:
            cursor.execute('ALTER TABLE users ADD COLUMN last_location TEXT')
        except:
            pass  # –ö–æ–ª–æ–Ω–∫–∞ —É–∂–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç
        
        try:
            cursor.execute('ALTER TABLE users ADD COLUMN last_status_change TIMESTAMP DEFAULT CURRENT_TIMESTAMP')
        except:
            pass  # –ö–æ–ª–æ–Ω–∫–∞ —É–∂–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç
        
        # –¢–∞–±–ª–∏—Ü–∞ –∑–∞–ø–∏—Å–µ–π
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                action TEXT NOT NULL,
                location TEXT NOT NULL,
                timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                comment TEXT,
                FOREIGN KEY (user_id) REFERENCES users (id)
            )
        ''')
        
        # –¢–∞–±–ª–∏—Ü–∞ –∞–¥–º–∏–Ω–æ–≤ —Å –ø—Ä–∞–≤–∞–º–∏
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS admins (
                id INTEGER PRIMARY KEY,
                username TEXT UNIQUE,
                permissions TEXT,
                appointed_by INTEGER,
                appointed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (appointed_by) REFERENCES users (id)
            )
        ''')
        
        # –î–æ–±–∞–≤–ª—è–µ–º –≥–ª–∞–≤–Ω–æ–≥–æ –∞–¥–º–∏–Ω–∞ –µ—Å–ª–∏ –µ–≥–æ –Ω–µ—Ç
        if MAIN_ADMIN_ID:
            cursor.execute('''
                INSERT OR IGNORE INTO users (id, username, full_name, is_admin)
                VALUES (?, ?, ?, ?)
            ''', (MAIN_ADMIN_ID, 'main_admin', '–ì–ª–∞–≤–Ω—ã–π –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä', True))
            
            cursor.execute('''
                INSERT OR IGNORE INTO admins (id, username, permissions, appointed_by)
                VALUES (?, ?, ?, ?)
            ''', (MAIN_ADMIN_ID, 'main_admin', 'all', MAIN_ADMIN_ID))
        
        conn.commit()
        conn.close()
    
    def add_user(self, user_id: int, username: str, full_name: str) -> bool:
        """–î–æ–±–∞–≤–ª–µ–Ω–∏–µ –Ω–æ–≤–æ–≥–æ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è"""
        try:
            conn = sqlite3.connect(self.db_name)
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT OR REPLACE INTO users (id, username, full_name)
                VALUES (?, ?, ?)
            ''', (user_id, username, full_name))
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"–û—à–∏–±–∫–∞ –¥–æ–±–∞–≤–ª–µ–Ω–∏—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è: {e}")
            return False
    
    def get_user(self, user_id: int) -> Optional[Dict]:
        """–ü–æ–ª—É—á–µ–Ω–∏–µ –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏–∏ –æ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT id, username, full_name, is_admin FROM users WHERE id = ?
        ''', (user_id,))
        
        result = cursor.fetchone()
        conn.close()
        
        if result:
            return {
                'id': result[0],
                'username': result[1],
                'full_name': result[2],
                'is_admin': bool(result[3])
            }
        return None
    
    def is_admin(self, user_id: int) -> bool:
        """–ü—Ä–æ–≤–µ—Ä–∫–∞ —è–≤–ª—è–µ—Ç—Å—è –ª–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å –∞–¥–º–∏–Ω–æ–º"""
        user = self.get_user(user_id)
        return user['is_admin'] if user else False
    
    def is_main_admin(self, user_id: int) -> bool:
        """–ü—Ä–æ–≤–µ—Ä–∫–∞ —è–≤–ª—è–µ—Ç—Å—è –ª–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å –≥–ª–∞–≤–Ω—ã–º –∞–¥–º–∏–Ω–æ–º"""
        return user_id == MAIN_ADMIN_ID
    
    def add_record(self, user_id: int, action: str, location: str, comment: str = None) -> bool:
        """–î–æ–±–∞–≤–ª–µ–Ω–∏–µ –∑–∞–ø–∏—Å–∏ –æ –≤—ã—Ö–æ–¥–µ/–ø—Ä–∏—Ö–æ–¥–µ"""
        try:
            conn = sqlite3.connect(self.db_name)
            cursor = conn.cursor()
            
            # –î–æ–±–∞–≤–ª—è–µ–º –∑–∞–ø–∏—Å—å
            cursor.execute('''
                INSERT INTO records (user_id, action, location, comment)
                VALUES (?, ?, ?, ?)
            ''', (user_id, action, location, comment))
            
            # –û–±–Ω–æ–≤–ª—è–µ–º —Å—Ç–∞—Ç—É—Å –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è
            new_status = '–≤–Ω–µ_—á–∞—Å—Ç–∏' if action == '—É–±—ã–ª' else '–≤_—á–∞—Å—Ç–∏'
            cursor.execute('''
                UPDATE users 
                SET status = ?, last_location = ?, last_status_change = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (new_status, location, user_id))
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"–û—à–∏–±–∫–∞ –¥–æ–±–∞–≤–ª–µ–Ω–∏—è –∑–∞–ø–∏—Å–∏: {e}")
            return False
    
    def get_user_records(self, user_id: int, limit: int = 10) -> List[Dict]:
        """–ü–æ–ª—É—á–µ–Ω–∏–µ –∑–∞–ø–∏—Å–µ–π –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT id, action, location, timestamp, comment
            FROM records 
            WHERE user_id = ?
            ORDER BY timestamp DESC
            LIMIT ?
        ''', (user_id, limit))
        
        results = cursor.fetchall()
        conn.close()
        
        return [
            {
                'id': row[0],
                'action': row[1],
                'location': row[2],
                'timestamp': row[3],
                'comment': row[4]
            }
            for row in results
        ]
    
    def get_all_records(self, days: int = 30) -> List[Dict]:
        """–ü–æ–ª—É—á–µ–Ω–∏–µ –≤—Å–µ—Ö –∑–∞–ø–∏—Å–µ–π –∑–∞ –ø–µ—Ä–∏–æ–¥"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT r.id, u.full_name, r.action, r.location, r.timestamp, r.comment
            FROM records r
            JOIN users u ON r.user_id = u.id
            WHERE r.timestamp >= datetime('now', '-{} days')
            ORDER BY r.timestamp DESC
        '''.format(days))
        
        results = cursor.fetchall()
        conn.close()
        
        return [
            {
                'id': row[0],
                'full_name': row[1],
                'action': row[2],
                'location': row[3],
                'timestamp': row[4],
                'comment': row[5]
            }
            for row in results
        ]
    
    def get_statistics(self, days: int = 30) -> Dict:
        """–ü–æ–ª—É—á–µ–Ω–∏–µ —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∏"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        
        # –û–±—â–∞—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞
        cursor.execute('''
            SELECT COUNT(*) FROM records 
            WHERE timestamp >= datetime('now', '-{} days')
        '''.format(days))
        total_records = cursor.fetchone()[0]
        
        # –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –ø–æ –¥–µ–π—Å—Ç–≤–∏—è–º
        cursor.execute('''
            SELECT action, COUNT(*) FROM records 
            WHERE timestamp >= datetime('now', '-{} days')
            GROUP BY action
        '''.format(days))
        action_stats = dict(cursor.fetchall())
        
        # –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –ø–æ –ª–æ–∫–∞—Ü–∏—è–º
        cursor.execute('''
            SELECT location, COUNT(*) FROM records 
            WHERE timestamp >= datetime('now', '-{} days')
            GROUP BY location
            ORDER BY COUNT(*) DESC
        '''.format(days))
        location_stats = dict(cursor.fetchall())
        
        # –ê–∫—Ç–∏–≤–Ω—ã–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–∏
        cursor.execute('''
            SELECT COUNT(DISTINCT user_id) FROM records 
            WHERE timestamp >= datetime('now', '-{} days')
        '''.format(days))
        active_users = cursor.fetchone()[0]
        
        conn.close()
        
        return {
            'total_records': total_records,
            'action_stats': action_stats,
            'location_stats': location_stats,
            'active_users': active_users
        }
    
    def export_to_excel(self, days: int = 30) -> str:
        """–≠–∫—Å–ø–æ—Ä—Ç –¥–∞–Ω–Ω—ã—Ö –≤ Excel"""
        records = self.get_all_records(days)
        
        if not records:
            return None
        
        df = pd.DataFrame(records)
        df['timestamp'] = pd.to_datetime(df['timestamp'])
        df = df.sort_values('timestamp', ascending=False)
        
        # –°–æ–∑–¥–∞–µ–º Excel —Ñ–∞–π–ª —Å —Ü–≤–µ—Ç–æ–≤–æ–π —Å—Ö–µ–º–æ–π
        with pd.ExcelWriter(EXPORT_FILENAME, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='–ó–∞–ø–∏—Å–∏', index=False)
            
            # –ü–æ–ª—É—á–∞–µ–º —Ä–∞–±–æ—á—É—é –∫–Ω–∏–≥—É –¥–ª—è —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏—è
            workbook = writer.book
            worksheet = writer.sheets['–ó–∞–ø–∏—Å–∏']
            
            # –ü—Ä–∏–º–µ–Ω—è–µ–º —Ü–≤–µ—Ç–æ–≤—É—é —Å—Ö–µ–º—É
            for row in range(2, len(df) + 2):  # –ù–∞—á–∏–Ω–∞–µ–º —Å 2 (–ø–æ—Å–ª–µ –∑–∞–≥–æ–ª–æ–≤–∫–æ–≤)
                action = worksheet.cell(row=row, column=3).value  # –ö–æ–ª–æ–Ω–∫–∞ action
                if action == '—É–±—ã–ª':
                    worksheet.cell(row=row, column=3).fill = openpyxl.styles.PatternFill(
                        start_color='FF0000', end_color='FF0000', fill_type='solid'
                    )
                elif action == '–ø—Ä–∏–±—ã–ª':
                    worksheet.cell(row=row, column=3).fill = openpyxl.styles.PatternFill(
                        start_color='00FF00', end_color='00FF00', fill_type='solid'
                    )
        
        return EXPORT_FILENAME
    
    def cleanup_old_records(self, months: int = 6):
        """–û—á–∏—Å—Ç–∫–∞ —Å—Ç–∞—Ä—ã—Ö –∑–∞–ø–∏—Å–µ–π (—Å—Ç–∞—Ä—à–µ 6 –º–µ—Å—è—Ü–µ–≤)"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        
        cursor.execute('''
            DELETE FROM records 
            WHERE timestamp < datetime('now', '-{} months')
        '''.format(months))
        
        deleted_count = cursor.rowcount
        conn.commit()
        conn.close()
        
        return deleted_count
    
    def get_all_admins(self) -> List[Dict]:
        """–ü–æ–ª—É—á–µ–Ω–∏–µ —Å–ø–∏—Å–∫–∞ –≤—Å–µ—Ö –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–æ–≤"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT u.id, u.username, u.full_name, a.permissions
            FROM users u
            JOIN admins a ON u.id = a.id
            WHERE u.is_admin = TRUE
            ORDER BY u.full_name
        ''')
        
        results = cursor.fetchall()
        conn.close()
        
        return [
            {
                'id': row[0],
                'username': row[1],
                'full_name': row[2],
                'permissions': row[3]
            }
            for row in results
        ]
    
    def add_admin(self, user_id: int, permissions: str = "basic") -> bool:
        """–î–æ–±–∞–≤–ª–µ–Ω–∏–µ –Ω–æ–≤–æ–≥–æ –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–∞"""
        try:
            conn = sqlite3.connect(self.db_name)
            cursor = conn.cursor()
            
            # –û–±–Ω–æ–≤–ª—è–µ–º —Å—Ç–∞—Ç—É—Å –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è
            cursor.execute('''
                UPDATE users SET is_admin = TRUE WHERE id = ?
            ''', (user_id,))
            
            # –î–æ–±–∞–≤–ª—è–µ–º –≤ —Ç–∞–±–ª–∏—Ü—É –∞–¥–º–∏–Ω–æ–≤
            user = self.get_user(user_id)
            if user:
                cursor.execute('''
                    INSERT OR REPLACE INTO admins (id, username, permissions, appointed_by)
                    VALUES (?, ?, ?, ?)
                ''', (user_id, user['username'], permissions, MAIN_ADMIN_ID))
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"–û—à–∏–±–∫–∞ –¥–æ–±–∞–≤–ª–µ–Ω–∏—è –∞–¥–º–∏–Ω–∞: {e}")
            return False
    
    def remove_admin(self, user_id: int) -> bool:
        """–£–¥–∞–ª–µ–Ω–∏–µ –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–∞"""
        try:
            conn = sqlite3.connect(self.db_name)
            cursor = conn.cursor()
            
            # –ù–µ–ª—å–∑—è —É–¥–∞–ª–∏—Ç—å –≥–ª–∞–≤–Ω–æ–≥–æ –∞–¥–º–∏–Ω–∞
            if user_id == MAIN_ADMIN_ID:
                return False
            
            # –û–±–Ω–æ–≤–ª—è–µ–º —Å—Ç–∞—Ç—É—Å –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è
            cursor.execute('''
                UPDATE users SET is_admin = FALSE WHERE id = ?
            ''', (user_id,))
            
            # –£–¥–∞–ª—è–µ–º –∏–∑ —Ç–∞–±–ª–∏—Ü—ã –∞–¥–º–∏–Ω–æ–≤
            cursor.execute('''
                DELETE FROM admins WHERE id = ?
            ''', (user_id,))
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"–û—à–∏–±–∫–∞ —É–¥–∞–ª–µ–Ω–∏—è –∞–¥–º–∏–Ω–∞: {e}")
            return False
    
    def get_user_by_id(self, user_id: int) -> Optional[Dict]:
        """–ü–æ–ª—É—á–µ–Ω–∏–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –ø–æ ID"""
        return self.get_user(user_id)
    
    def get_soldiers_by_status(self, status: str) -> List[Dict]:
        """–ü–æ–ª—É—á–µ–Ω–∏–µ –±–æ–π—Ü–æ–≤ –ø–æ —Å—Ç–∞—Ç—É—Å—É"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT id, username, full_name, status, last_location, last_status_change
            FROM users 
            WHERE status = ? AND is_admin = FALSE
            ORDER BY full_name
        ''', (status,))
        
        results = cursor.fetchall()
        conn.close()
        
        return [
            {
                'id': row[0],
                'username': row[1],
                'full_name': row[2],
                'status': row[3],
                'last_location': row[4],
                'last_status_change': row[5]
            }
            for row in results
        ]
    
    def mark_all_arrived(self) -> int:
        """–û—Ç–º–µ—Ç–∏—Ç—å –≤—Å–µ—Ö –±–æ–π—Ü–æ–≤ –ø—Ä–∏–±—ã–≤—à–∏–º–∏"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        
        cursor.execute('''
            UPDATE users 
            SET status = '–≤_—á–∞—Å—Ç–∏', last_status_change = CURRENT_TIMESTAMP
            WHERE is_admin = FALSE AND status = '–≤–Ω–µ_—á–∞—Å—Ç–∏'
        ''')
        
        updated_count = cursor.rowcount
        
        # –î–æ–±–∞–≤–ª—è–µ–º –∑–∞–ø–∏—Å–∏ –æ –ø—Ä–∏–±—ã—Ç–∏–∏
        cursor.execute('''
            SELECT id FROM users 
            WHERE is_admin = FALSE AND status = '–≤_—á–∞—Å—Ç–∏'
        ''')
        
        user_ids = [row[0] for row in cursor.fetchall()]
        
        for user_id in user_ids:
            cursor.execute('''
                INSERT INTO records (user_id, action, location, comment)
                VALUES (?, '–ø—Ä–∏–±—ã–ª', 'üè† –ß–∞—Å—Ç—å', '–ú–∞—Å—Å–æ–≤–∞—è –æ—Ç–º–µ—Ç–∫–∞')
            ''', (user_id,))
        
        conn.commit()
        conn.close()
        
        return updated_count
    
    def clear_all_data(self) -> bool:
        """–û—á–∏—Å—Ç–∏—Ç—å –≤—Å–µ –¥–∞–Ω–Ω—ã–µ (–æ–ø–∞—Å–Ω–∞—è –æ–ø–µ—Ä–∞—Ü–∏—è)"""
        try:
            conn = sqlite3.connect(self.db_name)
            cursor = conn.cursor()
            
            # –°–æ—Ö—Ä–∞–Ω—è–µ–º –≥–ª–∞–≤–Ω–æ–≥–æ –∞–¥–º–∏–Ω–∞
            main_admin = None
            if MAIN_ADMIN_ID:
                cursor.execute('SELECT * FROM users WHERE id = ?', (MAIN_ADMIN_ID,))
                main_admin = cursor.fetchone()
            
            # –û—á–∏—â–∞–µ–º –≤—Å–µ —Ç–∞–±–ª–∏—Ü—ã
            cursor.execute('DELETE FROM records')
            cursor.execute('DELETE FROM admins WHERE id != ?', (MAIN_ADMIN_ID,))
            cursor.execute('DELETE FROM users WHERE id != ?', (MAIN_ADMIN_ID,))
            
            # –í–æ—Å—Å—Ç–∞–Ω–∞–≤–ª–∏–≤–∞–µ–º –≥–ª–∞–≤–Ω–æ–≥–æ –∞–¥–º–∏–Ω–∞ –µ—Å–ª–∏ –Ω—É–∂–Ω–æ
            if main_admin and MAIN_ADMIN_ID:
                cursor.execute('''
                    INSERT OR REPLACE INTO users (id, username, full_name, is_admin)
                    VALUES (?, ?, ?, ?)
                ''', main_admin)
                
                cursor.execute('''
                    INSERT OR REPLACE INTO admins (id, username, permissions, appointed_by)
                    VALUES (?, ?, ?, ?)
                ''', (MAIN_ADMIN_ID, 'main_admin', 'all', MAIN_ADMIN_ID))
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"–û—à–∏–±–∫–∞ –æ—á–∏—Å—Ç–∫–∏ –¥–∞–Ω–Ω—ã—Ö: {e}")
            return False
    
    def get_users_list(self, page: int = 1, per_page: int = 8) -> tuple:
        """–ü–æ–ª—É—á–µ–Ω–∏–µ —Å–ø–∏—Å–∫–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π —Å –ø–∞–≥–∏–Ω–∞—Ü–∏–µ–π"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        
        # –û–±—â–µ–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π
        cursor.execute('SELECT COUNT(*) FROM users WHERE is_admin = FALSE')
        total_users = cursor.fetchone()[0]
        
        # –ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–∏ –¥–ª—è —Ç–µ–∫—É—â–µ–π —Å—Ç—Ä–∞–Ω–∏—Ü—ã
        offset = (page - 1) * per_page
        cursor.execute('''
            SELECT id, username, full_name, status, last_location, last_status_change
            FROM users 
            WHERE is_admin = FALSE
            ORDER BY full_name
            LIMIT ? OFFSET ?
        ''', (per_page, offset))
        
        results = cursor.fetchall()
        conn.close()
        
        users = [
            {
                'id': row[0],
                'username': row[1],
                'full_name': row[2],
                'status': row[3],
                'last_location': row[4],
                'last_status_change': row[5]
            }
            for row in results
        ]
        
        total_pages = (total_users + per_page - 1) // per_page
        
        return users, page, total_pages