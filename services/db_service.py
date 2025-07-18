import sqlite3
import logging
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Any
import os

# Проверяем наличие необходимых библиотек
try:
    import pandas as pd
    import openpyxl
    EXPORT_AVAILABLE = True
except ImportError as e:
    logging.error(f"❌ Ошибка импорта библиотек для экспорта: {e}")
    logging.error("Установите зависимости: pip install pandas openpyxl")
    EXPORT_AVAILABLE = False
    pd = None

class DatabaseService:
    def __init__(self, db_path: str = "military_tracker.db"):
        self.db_path = db_path
        self.init_db()

    def init_db(self):
        """Инициализация базы данных"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.execute('''
                    CREATE TABLE IF NOT EXISTS users (
                        id INTEGER PRIMARY KEY,
                        username TEXT NOT NULL,
                        full_name TEXT NOT NULL,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        is_admin BOOLEAN DEFAULT FALSE
                    )
                ''')

                conn.execute('''
                    CREATE TABLE IF NOT EXISTS records (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        user_id INTEGER NOT NULL,
                        action TEXT NOT NULL,
                        location TEXT NOT NULL,
                        timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (user_id) REFERENCES users (id)
                    )
                ''')

                conn.execute('''
                    CREATE TABLE IF NOT EXISTS admins (
                        user_id INTEGER PRIMARY KEY,
                        added_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (user_id) REFERENCES users (id)
                    )
                ''')

                # Создаем индексы для улучшения производительности
                conn.execute('CREATE INDEX IF NOT EXISTS idx_records_user_id ON records (user_id)')
                conn.execute('CREATE INDEX IF NOT EXISTS idx_records_timestamp ON records (timestamp)')

                # Миграция: добавляем отсутствующую колонку added_at если её нет
                try:
                    conn.execute('ALTER TABLE admins ADD COLUMN added_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP')
                    pass  # Колонка успешно добавлена
                except sqlite3.OperationalError as e:
                    if "duplicate column name" not in str(e).lower():
                        logging.warning(f"⚠️ Ошибка добавления колонки added_at: {e}")

                # Ensure main admin is added
                self.ensure_main_admin(conn)

                conn.commit()
        except Exception as e:
            logging.error(f"Ошибка инициализации БД: {e}")

    def ensure_main_admin(self, conn: sqlite3.Connection):
        """Гарантирует, что главный админ добавлен в таблицу"""
        try:
            from config import MAIN_ADMIN_ID
            cursor = conn.cursor()
            cursor.execute(
                'INSERT OR IGNORE INTO admins (user_id, added_at) VALUES (?, CURRENT_TIMESTAMP)',
                (MAIN_ADMIN_ID,)
            )
            pass  # Главный админ добавлен
        except ImportError:
            logging.warning("⚠️ config.py не найден, главный админ не добавлен")
        except Exception as e:
            logging.error(f"❌ Ошибка добавления главного админа: {e}")

    def add_user(self, user_id: int, username: str, full_name: str) -> bool:
        """Добавить пользователя"""
        try:
            # Валидация входных данных
            if not isinstance(user_id, int) or user_id <= 0:
                logging.error(f"Некорректный user_id: {user_id}")
                return False

            if not username or len(username.strip()) == 0:
                logging.error("Пустое имя пользователя")
                return False

            if not full_name or len(full_name.strip()) < 3 or len(full_name.strip()) > 50:
                logging.error(f"Некорректное ФИО: {full_name}")
                return False

            username = username.strip()[:50]  # Ограничиваем длину
            full_name = full_name.strip()

            with sqlite3.connect(self.db_path) as conn:
                conn.execute(
                    'INSERT OR REPLACE INTO users (id, username, full_name) VALUES (?, ?, ?)',
                    (user_id, username, full_name)
                )
                conn.commit()
                return True
        except Exception as e:
            logging.error(f"Ошибка добавления пользователя: {e}")
            return False

    def get_user(self, user_id: int) -> Optional[Dict[str, Any]]:
        """Получить пользователя"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.execute(
                    'SELECT * FROM users WHERE id = ?',
                    (user_id,)
                )
                row = cursor.fetchone()
                return dict(row) if row else None
        except Exception as e:
            logging.error(f"Ошибка получения пользователя: {e}")
            return None

    def add_record(self, user_id: int, action: str, location: str) -> bool:
        """Добавить запись"""
        try:
            # Валидация входных данных
            if not isinstance(user_id, int) or user_id <= 0:
                logging.error(f"Некорректный user_id: {user_id}")
                return False

            if not action or action.strip() not in ['в части', 'не в части']:
                logging.error(f"Некорректное действие: {action}")
                return False

            if not location or len(location.strip()) < 1 or len(location.strip()) > 100:
                logging.error(f"Некорректная локация: {location}")
                return False

            # Проверяем, существует ли пользователь
            if not self.get_user(user_id):
                logging.error(f"Пользователь {user_id} не найден")
                return False

            action = action.strip()
            location = location.strip()

            # Мягкая защита от дублирования записей
            last_records = self.get_user_records(user_id, 1)
            if last_records:
                last_record = last_records[0]  # Первая запись (самая новая)
                last_action = last_record['action']
                last_timestamp = datetime.fromisoformat(last_record['timestamp'].replace('Z', '+00:00'))

                # Проверяем, было ли такое же действие в последние 3 секунды (уменьшили время)
                time_diff = (datetime.now() - last_timestamp.replace(tzinfo=None)).total_seconds()

                # Блокируем только явные дубли в течение 3 секунд
                if last_action == action and last_record['location'] == location and time_diff < 3:
                    logging.warning(f"Быстрое дублирование записи заблокировано для пользователя {user_id} (разница: {time_diff:.1f}с)")
                    return False

            with sqlite3.connect(self.db_path) as conn:
                conn.execute(
                    'INSERT INTO records (user_id, action, location) VALUES (?, ?, ?)',
                    (user_id, action, location)
                )
                conn.commit()
                return True
        except Exception as e:
            logging.error(f"Ошибка добавления записи: {e}")
            return False

    def get_user_records(self, user_id: int, limit: int = 10) -> List[Dict[str, Any]]:
        """Получить записи пользователя"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.execute(
                    'SELECT * FROM records WHERE user_id = ? ORDER BY timestamp DESC LIMIT ?',
                    (user_id, limit)
                )
                return [dict(row) for row in cursor.fetchall()]
        except Exception as e:
            logging.error(f"Ошибка получения записей пользователя: {e}")
            return []

    def get_all_records(self, days: int = 7, limit: int = 100) -> List[Dict[str, Any]]:
        """Получить все записи за период"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                since_date = datetime.now() - timedelta(days=days)
                cursor = conn.execute('''
                    SELECT r.*, u.full_name 
                    FROM records r
                    JOIN users u ON r.user_id = u.id
                    WHERE r.timestamp > ?
                    ORDER BY r.timestamp DESC
                    LIMIT ?
                ''', (since_date, limit))
                return [dict(row) for row in cursor.fetchall()]
        except Exception as e:
            logging.error(f"Ошибка получения всех записей: {e}")
            return []

    def get_records_paginated(self, page: int = 1, per_page: int = 10, days: int = 7, 
                            user_filter: str = None, location_filter: str = None) -> Dict[str, Any]:
        """Получить записи с пагинацией и фильтрами"""
        try:
            offset = (page - 1) * per_page
            since_date = datetime.now() - timedelta(days=days)

            # Базовый запрос
            base_query = '''
                SELECT r.*, u.full_name 
                FROM records r
                JOIN users u ON r.user_id = u.id
                WHERE r.timestamp > ?
            '''
            count_query = '''
                SELECT COUNT(*) as total
                FROM records r
                JOIN users u ON r.user_id = u.id
                WHERE r.timestamp > ?
            '''

            params = [since_date]

            # Добавляем фильтры
            if user_filter:
                base_query += ' AND u.full_name LIKE ?'
                count_query += ' AND u.full_name LIKE ?'
                params.append(f'%{user_filter}%')

            if location_filter:
                base_query += ' AND r.location LIKE ?'
                count_query += ' AND r.location LIKE ?'
                params.append(f'%{location_filter}%')

            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row

                # Получаем общее количество
                total_cursor = conn.execute(count_query, params)
                total_records = total_cursor.fetchone()['total']

                # Получаем записи для текущей страницы
                base_query += ' ORDER BY r.timestamp DESC LIMIT ? OFFSET ?'
                cursor = conn.execute(base_query, params + [per_page, offset])
                records = [dict(row) for row in cursor.fetchall()]

                total_pages = (total_records + per_page - 1) // per_page

                return {
                    'records': records,
                    'current_page': page,
                    'total_pages': total_pages,
                    'total_records': total_records,
                    'per_page': per_page,
                    'has_prev': page > 1,
                    'has_next': page < total_pages
                }

        except Exception as e:
            logging.error(f"Ошибка получения записей с пагинацией: {e}")
            return {
                'records': [],
                'current_page': 1,
                'total_pages': 0,
                'total_records': 0,
                'per_page': per_page,
                'has_prev': False,
                'has_next': False
            }

    def get_users_paginated(self, page: int = 1, per_page: int = 20, search: str = None) -> Dict[str, Any]:
        """Получить пользователей с пагинацией"""
        try:
            offset = (page - 1) * per_page

            base_query = 'SELECT * FROM users'
            count_query = 'SELECT COUNT(*) as total FROM users'
            params = []

            if search:
                base_query += ' WHERE full_name LIKE ? OR username LIKE ?'
                count_query += ' WHERE full_name LIKE ? OR username LIKE ?'
                params = [f'%{search}%', f'%{search}%']

            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row

                # Получаем общее количество
                total_cursor = conn.execute(count_query, params)
                total_users = total_cursor.fetchone()['total']

                # Получаем пользователей для текущей страницы
                base_query += ' ORDER BY full_name LIMIT ? OFFSET ?'
                cursor = conn.execute(base_query, params + [per_page, offset])
                users = [dict(row) for row in cursor.fetchall()]

                total_pages = (total_users + per_page - 1) // per_page

                return {
                    'users': users,
                    'current_page': page,
                    'total_pages': total_pages,
                    'total_users': total_users,
                    'per_page': per_page,
                    'has_prev': page > 1,
                    'has_next': page < total_pages
                }

        except Exception as e:
            logging.error(f"Ошибка получения пользователей с пагинацией: {e}")
            return {
                'users': [],
                'current_page': 1,
                'total_pages': 0,
                'total_users': 0,
                'per_page': per_page,
                'has_prev': False,
                'has_next': False
            }

    def get_current_status(self) -> Dict[str, Any]:
        """Получить текущий статус всех пользователей с группировкой по локациям"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row

                # Получаем всех пользователей
                users_cursor = conn.execute('SELECT id, full_name FROM users')
                all_users = users_cursor.fetchall()

                # Получаем последние действия каждого пользователя
                absent_users = []
                present_users = []

                # Группировка по локациям
                location_groups = {}

                for user in all_users:
                    last_record_cursor = conn.execute('''
                        SELECT action, location FROM records 
                        WHERE user_id = ? 
                        ORDER BY timestamp DESC 
                        LIMIT 1
                    ''', (user['id'],))
                    last_record = last_record_cursor.fetchone()

                    # Проверяем статус: "не в части" = отсутствует, "в части" = присутствует
                    if last_record and last_record['action'] == 'не в части':
                        location = last_record['location']
                        absent_users.append({
                            'name': user['full_name'],
                            'location': location
                        })

                        # Группируем отсутствующих по локациям
                        if location not in location_groups:
                            location_groups[location] = {'count': 0, 'names': []}
                        location_groups[location]['count'] += 1
                        location_groups[location]['names'].append(user['full_name'])

                    else:
                        # Если последнее действие "в части" или записей нет - считаем в части
                        present_users.append({
                            'name': user['full_name'],
                            'location': 'В части'
                        })

                        # Группируем присутствующих
                        if 'В части' not in location_groups:
                            location_groups['В части'] = {'count': 0, 'names': []}
                        location_groups['В части']['count'] += 1
                        location_groups['В части']['names'].append(user['full_name'])

                return {
                    'total': len(all_users),
                    'present': len(present_users),
                    'absent': len(absent_users),
                    'absent_list': absent_users,
                    'present_list': present_users,
                    'location_groups': location_groups
                }
        except Exception as e:
            logging.error(f"Ошибка получения статуса: {e}")
            return {'total': 0, 'present': 0, 'absent': 0, 'absent_list': []}

    def is_admin(self, user_id: int) -> bool:
        """Проверить права администратора"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.execute(
                    'SELECT 1 FROM admins WHERE user_id = ?',
                    (user_id,)
                )
                return cursor.fetchone() is not None
        except Exception as e:
            logging.error(f"Ошибка проверки прав админа: {e}")
            return False

    def add_admin(self, user_id: int) -> bool:
        """Добавить администратора"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.execute(
                    'INSERT OR IGNORE INTO admins (user_id) VALUES (?)',
                    (user_id,)
                )
                conn.commit()
                return True
        except Exception as e:
            logging.error(f"Ошибка добавления админа: {e}")
            return False

    def get_all_admins(self) -> List[Dict[str, Any]]:
        """Получить всех админов"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.execute('''
                    SELECT u.id, u.username, u.full_name, a.added_at
                    FROM admins a
                    JOIN users u ON a.user_id = u.id
                    ORDER BY u.full_name
                ''')
                return [dict(row) for row in cursor.fetchall()]
        except Exception as e:
            logging.error(f"Ошибка получения админов: {e}")
            # Fallback запрос без added_at если колонки нет
            try:
                cursor = conn.execute('''
                    SELECT u.id, u.username, u.full_name
                    FROM admins a
                    JOIN users u ON a.user_id = u.id
                    ORDER BY u.full_name
                ''')
                return [dict(row) for row in cursor.fetchall()]
            except Exception as e2:
                logging.error(f"Ошибка fallback запроса админов: {e2}")
                return []

    def get_all_users(self) -> List[Dict[str, Any]]:
        """Получить всех пользователей"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.execute('SELECT * FROM users ORDER BY full_name')
                return [dict(row) for row in cursor.fetchall()]
        except Exception as e:
            logging.error(f"Ошибка получения пользователей: {e}")
            return []

    def export_to_excel(self, days: int = 30) -> Optional[str]:
        """Экспорт данных в Excel"""
        if not EXPORT_AVAILABLE:
            logging.error("❌ Библиотеки для экспорта недоступны")
            return None

        try:
            records = self.get_all_records(days=days, limit=10000)
            if not records:
                logging.warning("Нет записей для экспорта")
                return None

            return self.export_records_to_excel(records, f"за последние {days} дней")
        except Exception as e:
            logging.error(f"Ошибка экспорта в Excel: {e}")
            return None

    def export_records_to_excel(self, records: List[Dict[str, Any]], period_desc: str = "") -> Optional[str]:
        """Экспорт списка записей в Excel с форматированием"""
        if not EXPORT_AVAILABLE:
            logging.error("❌ Библиотеки для экспорта недоступны")
            return None

        try:
            if not records:
                logging.warning("Нет записей для экспорта")
                return None

            # Создаем DataFrame
            df = pd.DataFrame(records)

            # Форматируем данные - сортируем по времени (хронологический порядок)
            df['timestamp'] = pd.to_datetime(df['timestamp'])
            df = df.sort_values('timestamp', ascending=True)

            # Преобразуем действия для корректного отображения
            df['action'] = df['action'].replace({
                'в части': 'прибыл', 
                'не в части': 'убыл'
            })

            # Переименовываем колонки
            df = df.rename(columns={
                'full_name': 'ФИО',
                'action': 'Действие', 
                'location': 'Локация',
                'timestamp': 'Дата_Время'
            })

            # Убираем эмодзи из локаций
            df['Локация'] = df['Локация'].str.replace(r'[^\w\s\-\.\,\(\)]', '', regex=True).str.strip()

            # Создаем отдельные столбцы для даты и времени
            df['Дата'] = df['Дата_Время'].dt.strftime('%d.%m.%Y')
            df['Время'] = df['Дата_Время'].dt.strftime('%H:%M:%S')

            # Выбираем нужные колонки в правильном порядке
            df = df[['ФИО', 'Действие', 'Локация', 'Дата', 'Время']]

            # Создаем уникальное имя файла
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"military_records_{timestamp}.xlsx"

            logging.info(f"Создаем Excel файл: {filename}")

            # Создаем Excel файл с улучшенным форматированием
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Записи', index=False)
                logging.info("DataFrame записан в Excel")

                # Получаем рабочий лист
                worksheet = writer.sheets['Записи']

                try:
                    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

                    # Определяем стили
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True, size=12)

                    arrived_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Светло-зеленый
                    departed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Светло-красный

                    border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

                    # Форматируем заголовки
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        try:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        except Exception as style_e:
                            logging.warning(f"Не удалось применить выравнивание: {style_e}")
                        cell.border = border

                    # Форматируем данные с цветовой заливкой
                    for row_num, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                        try:
                            action_cell = row[1]  # Колонка "Действие"

                            # Применяем цветовую заливку в зависимости от действия
                            if action_cell.value == "прибыл":
                                for cell in row:
                                    cell.fill = arrived_fill
                                    cell.border = border
                            elif action_cell.value == "убыл":
                                for cell in row:
                                    cell.fill = departed_fill
                                    cell.border = border
                            else:
                                for cell in row:
                                    cell.border = border
                        except Exception as row_e:
                            logging.warning(f"Ошибка форматирования строки {row_num}: {row_e}")

                    # Автоматически подгоняем ширину колонок
                    column_widths = {
                        'A': 20,  # ФИО - минимум 20
                        'B': 12,  # Действие - минимум 12
                        'C': 15,  # Локация - минимум 15
                        'D': 12,  # Дата - фиксированная ширина
                        'E': 10   # Время - фиксированная ширина
                    }

                    # Определяем максимальную ширину для каждой колонки
                    for row in worksheet.iter_rows():
                        for cell in row:
                            column_letter = cell.column_letter
                            if column_letter in column_widths:
                                try:
                                    cell_length = len(str(cell.value)) if cell.value else 0
                                    if cell_length > column_widths[column_letter]:
                                        column_widths[column_letter] = min(cell_length + 3, 50)  # Ограничиваем максимум
                                except Exception:
                                    pass

                    # Устанавливаем ширину колонок
                    for col_letter, width in column_widths.items():
                        worksheet.column_dimensions[col_letter].width = width

                    # Устанавливаем высоту строк
                    for row in worksheet.iter_rows():
                        worksheet.row_dimensions[row[0].row].height = 20

                    logging.info("Форматирование Excel файла завершено")

                except ImportError as import_e:
                    logging.warning(f"Опции стиля недоступны: {import_e}")
                except Exception as format_e:
                    logging.warning(f"Ошибка форматирования Excel: {format_e}")

            logging.info(f"Excel файл создан успешно: {filename}")
            return filename

        except Exception as e:
            logging.error(f"Ошибка создания Excel файла: {e}")
            import traceback
            logging.error(f"Трассировка: {traceback.format_exc()}")
            return None

    def export_to_csv(self, days: int = 30) -> str:
        """Экспорт записей в CSV файл"""
        try:
            import pandas as pd
            from datetime import datetime, timedelta

            # Получаем записи за период
            records = self.get_all_records(days=days, limit=10000)

            if not records:
                return None

            # Создаем DataFrame
            df = pd.DataFrame(records)

            # Форматируем временные метки
            df['timestamp'] = pd.to_datetime(df['timestamp'])
            df['date'] = df['timestamp'].dt.strftime('%d.%m.%Y')
            df['time'] = df['timestamp'].dt.strftime('%H:%M:%S')

            # Переупорядочиваем колонки
            df = df[['date', 'time', 'full_name', 'action', 'location']]
            df.columns = ['Дата', 'Время', 'ФИО', 'Действие', 'Локация']

            # Создаем имя файла
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"export_{timestamp}.csv"

            # Сохраняем файл
            df.to_csv(filename, index=False, encoding='utf-8-sig')

            return filename

        except Exception as e:
            logging.error(f"Ошибка экспорта CSV: {e}")
            return None

    def export_records_to_excel(self, records: list, period_desc: str) -> str:
        """Экспорт записей в Excel файл"""
        try:
            import pandas as pd
            from datetime import datetime

            if not records:
                return None

            # Создаем DataFrame
            df = pd.DataFrame(records)

            # Форматируем временные метки
            df['timestamp'] = pd.to_datetime(df['timestamp'])
            df['date'] = df['timestamp'].dt.strftime('%d.%m.%Y')
            df['time'] = df['timestamp'].dt.strftime('%H:%M:%S')

            # Переупорядочиваем колонки
            df = df[['date', 'time', 'full_name', 'action', 'location']]
            df.columns = ['Дата', 'Время', 'ФИО', 'Действие', 'Локация']

            # Создаем имя файла
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"export_excel_{timestamp}.xlsx"

            # Сохраняем в Excel
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Записи', index=False)

                # Добавляем информационный лист
                info_df = pd.DataFrame({
                    'Параметр': ['Период', 'Дата создания', 'Всего записей'],
                    'Значение': [period_desc, datetime.now().strftime('%d.%m.%Y %H:%M:%S'), len(records)]
                })
                info_df.to_excel(writer, sheet_name='Информация', index=False)

            return filename

        except Exception as e:
            logging.error(f"Ошибка экспорта Excel: {e}")
            return None

    def get_records_today(self) -> list:
        """Получить записи за сегодня"""
        try:
            from datetime import datetime, timedelta

            today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            tomorrow = today + timedelta(days=1)

            cursor = self.conn.execute("""
                SELECT * FROM records 
                WHERE timestamp >= ? AND timestamp < ?
                ORDER BY timestamp DESC
            """, (today.isoformat(), tomorrow.isoformat()))

            records = []
            for row in cursor.fetchall():
                records.append({
                    'id': row[0],
                    'user_id': row[1],
                    'full_name': row[2],
                    'action': row[3],
                    'location': row[4],
                    'timestamp': row[5]
                })

            return records

        except Exception as e:
            logging.error(f"Ошибка получения записей за сегодня: {e}")
            return []

    def get_records_yesterday(self) -> list:
        """Получить записи за вчера"""
        try:
            from datetime import datetime, timedelta

            yesterday = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=1)
            today = yesterday + timedelta(days=1)

            cursor = self.conn.execute("""
                SELECT * FROM records 
                WHERE timestamp >= ? AND timestamp < ?
                ORDER BY timestamp DESC
            """, (yesterday.isoformat(), today.isoformat()))

            records = []
            for row in cursor.fetchall():
                records.append({
                    'id': row[0],
                    'user_id': row[1],
                    'full_name': row[2],
                    'action': row[3],
                    'location': row[4],
                    'timestamp': row[5]
                })

            return records

        except Exception as e:
            logging.error(f"Ошибка получения записей за вчера: {e}")
            return []

    def get_records_by_date(self, date_str: str) -> List[Dict[str, Any]]:
        """Получить записи за конкретную дату"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                # Используем LIKE для более надежного поиска по дате
                cursor = conn.execute('''
                    SELECT r.*, u.full_name 
                    FROM records r
                    JOIN users u ON r.user_id = u.id
                    WHERE DATE(r.timestamp) = ? OR r.timestamp LIKE ?
                    ORDER BY r.timestamp ASC
                ''', (date_str, f"{date_str}%"))
                records = [dict(row) for row in cursor.fetchall()]
                logging.info(f"Найдено записей за {date_str}: {len(records)}")
                return records
        except Exception as e:
            logging.error(f"Ошибка получения записей по дате {date_str}: {e}")
            return []

    def get_records_today(self) -> List[Dict[str, Any]]:
        """Получить записи за сегодня"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                today = datetime.now().date()
                cursor = conn.execute('''
                    SELECT r.*, u.full_name 
                    FROM records r
                    JOIN users u ON r.user_id = u.id
                    WHERE DATE(r.timestamp) = ?
                    ORDER BY r.timestamp ASC
                ''', (str(today),))
                records = [dict(row) for row in cursor.fetchall()]
                logging.info(f"Найдено записей за сегодня ({today}): {len(records)}")
                return records
        except Exception as e:
            logging.error(f"Ошибка получения записей за сегодня: {e}")
            return []

    def get_records_yesterday(self) -> List[Dict[str, Any]]:
        """Получить записи за вчера"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                yesterday = (datetime.now() - timedelta(days=1)).date()
                cursor = conn.execute('''
                    SELECT r.*, u.full_name 
                    FROM records r
                    JOIN users u ON r.user_id = u.id
                    WHERE DATE(r.timestamp) = ?
                    ORDER BY r.timestamp ASC
                ''', (str(yesterday),))
                records = [dict(row) for row in cursor.fetchall()]
                logging.info(f"Найдено записей за вчера ({yesterday}): {len(records)}")
                return records
        except Exception as e:
            logging.error(f"Ошибка получения записей за вчера: {e}")
            return []

    def cleanup_old_records(self, days: int = 180) -> int:
        """Очистка старых записей"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cutoff_date = datetime.now() - timedelta(days=days)
                cursor = conn.execute(
                    'DELETE FROM records WHERE timestamp < ?',
                    (cutoff_date,)
                )
                deleted_count = cursor.rowcount
                conn.commit()
                return deleted_count
        except Exception as e:
            logging.error(f"Ошибка очистки записей: {e}")
            return 0

    def clear_all_records(self) -> int:
        """Удалить все записи из системы"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.execute("DELETE FROM records")
                deleted_count = cursor.rowcount
                conn.commit()
                return deleted_count
        except Exception as e:
            logging.error(f"Ошибка при очистке всех записей: {e}")
            return 0

    def cleanup_all_records(self) -> int:
        """Альтернативный метод для полной очистки записей"""
        return self.clear_all_records()

    def clear_all_data(self) -> int:
        """Полная очистка всех данных системы"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                # Подсчитываем общее количество записей перед удалением
                total_records = 0

                # Считаем записи в каждой таблице
                cursor = conn.execute("SELECT COUNT(*) FROM records")
                records_count = cursor.fetchone()[0]
                total_records += records_count

                cursor = conn.execute("SELECT COUNT(*) FROM users")
                users_count = cursor.fetchone()[0]
                total_records += users_count

                cursor = conn.execute("SELECT COUNT(*) FROM admins")
                admins_count = cursor.fetchone()[0]
                total_records += admins_count

                # Удаляем все данные из всех таблиц
                conn.execute("DELETE FROM records")
                conn.execute("DELETE FROM users")
                conn.execute("DELETE FROM admins")

                # Сбрасываем автоинкремент
                conn.execute("DELETE FROM sqlite_sequence WHERE name IN ('records', 'users', 'admins')")

                conn.commit()
                logging.info(f"База данных полностью очищена. Удалено записей: {total_records}")
                return total_records

        except Exception as e:
            logging.error(f"Ошибка при полной очистке БД: {e}")
            return 0

    def full_database_reset(self):
        """Полная очистка базы данных"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                # Удаляем все данные из всех таблиц
                conn.execute("DELETE FROM records")
                conn.execute("DELETE FROM users")
                conn.execute("DELETE FROM admins")

                # Сбрасываем автоинкремент
                conn.execute("DELETE FROM sqlite_sequence WHERE name IN ('records', 'users', 'admins')")

                conn.commit()
                logging.info("База данных полностью очищена")

        except Exception as e:
            logging.error(f"Ошибка при полной очистке БД: {e}")
            raise

    def optimize_database(self):
        """Оптимизация базы данных"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.execute("VACUUM")
                conn.execute("ANALYZE")
                conn.commit()
                logging.info("База данных оптимизирована")
        except Exception as e:
            logging.error(f"Ошибка при оптимизации БД: {e}")

    def get_database_stats(self) -> dict:
        """Получить статистику базы данных"""
        try:
            stats = {}
            with sqlite3.connect(self.db_path) as conn:
                # Количество записей в таблицах
                stats['users'] = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
                stats['records'] = conn.execute("SELECT COUNT(*) FROM records").fetchone()[0]
                stats['admins'] = conn.execute("SELECT COUNT(*) FROM admins").fetchone()[0]

                # Размер базы данных
                import os
                if os.path.exists(self.db_path):
                    size_bytes = os.path.getsize(self.db_path)
                    stats['size'] = f"{size_bytes / (1024 * 1024):.2f} МБ"
                else:
                    stats['size'] = "0 МБ"

                # Последняя активность
                try:
                    cursor = conn.execute("SELECT MAX(timestamp) FROM records")
                    last_activity = cursor.fetchone()[0]
                    if last_activity:
                        last_time = datetime.fromisoformat(last_activity.replace('Z', '+00:00'))
                        stats['last_activity'] = last_time.strftime('%d.%m.%Y %H:%M')
                    else:
                        stats['last_activity'] = "Нет активности"
                except:
                    stats['last_activity'] = "Неизвестно"

            return stats

        except Exception as e:
            logging.error(f"Ошибка при получении статистики БД: {e}")
            return {
                'users': 0,
                'records': 0, 
                'admins': 0,
                'size': '0 МБ',
                'last_activity': 'Ошибка'
            }

    def remove_admin(self, user_id: int) -> bool:
        """Удалить администратора"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.execute("DELETE FROM admins WHERE user_id = ?", (user_id,))
                conn.commit()
                return cursor.rowcount > 0
        except Exception as e:
            logging.error(f"Ошибка при удалении администратора: {e}")
            return False

    def delete_admin(self, user_id: int) -> bool:
        """Альтернативный метод удаления администратора"""
        return self.remove_admin(user_id)